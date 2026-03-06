"""
Step 2: Prepare All Files for Subagent Examination

Processes every file in the working directory into a format subagents can read:
- PDFs → rendered page-by-page to PNG (via PyMuPDF)
- Images → converted to PNG, resized if >2000px (via Pillow)
- Text files → copied as .txt (built-in)
- HTML/web files → copied as .txt (built-in)
- RTF → copied as .txt (noisy but readable by subagent)
- Spreadsheets → converted to CSV text (via openpyxl/xlrd)
- DOCX → text extracted (via python-docx)
- DOC → text extracted (via doc2txt)
- ZIP → extracted, then contents processed like everything else

Only installs packages that are actually needed based on file inventory.

Usage:
    python prepare.py                     # scan cwd
    python prepare.py /path/to/tax/docs   # scan specified directory

Output directory: <base>/.tmp_prepared/
"""

import os
import re
import shutil
import subprocess
import sys
import zipfile

base = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
outdir = os.path.join(base, '.tmp_prepared')
os.makedirs(outdir, exist_ok=True)

# --- File type categories ---

PDF_EXTS = {'.pdf'}
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.gif', '.tiff', '.tif', '.bmp',
              '.webp', '.avif', '.heic', '.heif'}
TEXT_EXTS = {'.txt', '.csv', '.md', '.json', '.log', '.rtf'}
HTML_EXTS = {'.html', '.htm', '.mhtml', '.mht', '.wbl'}
SPREADSHEET_EXTS = {'.xlsx', '.xls'}
DOCX_EXTS = {'.docx'}
DOC_EXTS = {'.doc'}
ZIP_EXTS = {'.zip'}

ALL_SUPPORTED = (PDF_EXTS | IMAGE_EXTS | TEXT_EXTS | HTML_EXTS |
                 SPREADSHEET_EXTS | DOCX_EXTS | DOC_EXTS | ZIP_EXTS)

# --- Collect all files ---

def collect_files(directory):
    """Recursively collect all files, skipping hidden directories."""
    files = []
    for root, dirs, filenames in os.walk(directory):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        for f in filenames:
            files.append(os.path.relpath(os.path.join(root, f), base))
    return files

all_files = collect_files(base)

# --- Extract ZIPs first ---

zip_files = [f for f in all_files if os.path.splitext(f)[1].lower() in ZIP_EXTS]
for zf in zip_files:
    fullpath = os.path.join(base, zf)
    extract_dir = os.path.join(base, '.tmp_zip_extracted',
                               os.path.splitext(os.path.basename(zf))[0])
    try:
        with zipfile.ZipFile(fullpath, 'r') as z:
            z.extractall(extract_dir)
        print(f'EXTRACTED ZIP | {zf} -> {os.path.relpath(extract_dir, base)}')
        # Add extracted files to the list
        for root, dirs, filenames in os.walk(extract_dir):
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for f in filenames:
                all_files.append(os.path.relpath(os.path.join(root, f), base))
    except Exception as e:
        print(f'ZIP ERROR | {zf} | {e}')

# Remove ZIPs from processing list (already extracted)
all_files = [f for f in all_files if os.path.splitext(f)[1].lower() not in ZIP_EXTS]

# --- Determine which packages are needed ---

found_exts = set()
for f in all_files:
    found_exts.add(os.path.splitext(f)[1].lower())

def try_import(module_name):
    try:
        __import__(module_name)
        return True
    except ImportError:
        return False

def install_package(pip_name):
    print(f'Installing {pip_name}...')
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pip_name],
                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

# Conditional installs
if found_exts & PDF_EXTS:
    if not try_import('fitz'):
        install_package('pymupdf')
    import fitz

if found_exts & IMAGE_EXTS:
    if not try_import('PIL'):
        install_package('Pillow')
    if found_exts & {'.heic', '.heif'}:
        if not try_import('pillow_heif'):
            install_package('pillow-heif')
        import pillow_heif
        pillow_heif.register_heif_opener()
    if found_exts & {'.avif'}:
        if not try_import('pillow_avif'):
            install_package('pillow-avif-plugin')
        import pillow_avif
    from PIL import Image

if found_exts & {'.xlsx'}:
    if not try_import('openpyxl'):
        install_package('openpyxl')
    import openpyxl

if found_exts & {'.xls'}:
    if not try_import('xlrd'):
        install_package('xlrd')
    import xlrd

if found_exts & DOCX_EXTS:
    if not try_import('docx'):
        install_package('python-docx')
    import docx

if found_exts & DOC_EXTS:
    if not try_import('doc2txt'):
        install_package('doc2txt')
    import doc2txt

# --- Helper functions ---

def sanitize_name(relpath, max_len=60):
    """Convert a relative path into a safe filename prefix."""
    parts = os.path.normpath(relpath).replace(os.sep, '_')
    return re.sub(r'[^a-zA-Z0-9_-]', '_', os.path.splitext(parts)[0])[:max_len]

def write_text_file(relpath, text, suffix=''):
    """Write a text file to the output directory."""
    name = sanitize_name(relpath)
    out_name = f'{name}{suffix}.txt'
    out_path = os.path.join(outdir, out_name)
    with open(out_path, 'w', encoding='utf-8', errors='replace') as f:
        f.write(text)
    return out_name

# --- Processing ---

total_outputs = 0
errors = []
manifest = []
unsupported = []

for relpath in sorted(all_files):
    fullpath = os.path.join(base, relpath)
    if not os.path.exists(fullpath):
        errors.append(f'MISSING: {relpath}')
        continue

    ext = os.path.splitext(relpath)[1].lower()

    try:
        # --- PDFs ---
        if ext in PDF_EXTS:
            doc = fitz.open(fullpath)
            prefix = sanitize_name(relpath)
            for i in range(len(doc)):
                pix = doc[i].get_pixmap(dpi=100)
                png_name = f'{prefix}_p{i+1:02d}.png'
                pix.save(os.path.join(outdir, png_name))
                manifest.append(f'{png_name} <- {relpath} (page {i+1})')
                total_outputs += 1
            doc.close()
            print(f'OK {len(doc):3d}p | {relpath}')

        # --- Images ---
        elif ext in IMAGE_EXTS:
            img = Image.open(fullpath)
            # Resize if either dimension exceeds 2000px
            max_dim = max(img.size)
            if max_dim > 2000:
                scale = 2000 / max_dim
                new_size = (int(img.size[0] * scale), int(img.size[1] * scale))
                img = img.resize(new_size, Image.LANCZOS)
            png_name = f'{sanitize_name(relpath)}.png'
            img.convert('RGB').save(os.path.join(outdir, png_name))
            manifest.append(f'{png_name} <- {relpath}')
            total_outputs += 1
            print(f'OK  img | {relpath}')

        # --- Text / HTML / RTF (already text, copy as-is) ---
        elif ext in TEXT_EXTS or ext in HTML_EXTS:
            with open(fullpath, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
            out_name = write_text_file(relpath, text)
            manifest.append(f'{out_name} <- {relpath}')
            total_outputs += 1
            print(f'OK  txt | {relpath}')

        # --- XLSX ---
        elif ext == '.xlsx':
            wb = openpyxl.load_workbook(fullpath, read_only=True, data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_text.append(f'=== Sheet: {sheet_name} ===')
                for row in ws.iter_rows(values_only=True):
                    all_text.append(','.join(str(c) if c is not None else '' for c in row))
            wb.close()
            out_name = write_text_file(relpath, '\n'.join(all_text))
            manifest.append(f'{out_name} <- {relpath}')
            total_outputs += 1
            print(f'OK xlsx | {relpath}')

        # --- XLS ---
        elif ext == '.xls':
            wb = xlrd.open_workbook(fullpath)
            all_text = []
            for sheet in wb.sheets():
                all_text.append(f'=== Sheet: {sheet.name} ===')
                for row_idx in range(sheet.nrows):
                    row = [str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)]
                    all_text.append(','.join(row))
            out_name = write_text_file(relpath, '\n'.join(all_text))
            manifest.append(f'{out_name} <- {relpath}')
            total_outputs += 1
            print(f'OK  xls | {relpath}')

        # --- DOCX ---
        elif ext in DOCX_EXTS:
            doc = docx.Document(fullpath)
            text = '\n'.join(p.text for p in doc.paragraphs)
            out_name = write_text_file(relpath, text)
            manifest.append(f'{out_name} <- {relpath}')
            total_outputs += 1
            print(f'OK docx | {relpath}')

        # --- DOC ---
        elif ext in DOC_EXTS:
            result = doc2txt.extract(fullpath)
            out_name = write_text_file(relpath, result)
            manifest.append(f'{out_name} <- {relpath}')
            total_outputs += 1
            print(f'OK  doc | {relpath}')

        # --- Unsupported ---
        else:
            unsupported.append(relpath)

    except Exception as e:
        errors.append(f'ERROR: {relpath} | {e}')

# --- Write manifest ---
manifest_path = os.path.join(outdir, 'manifest.txt')
with open(manifest_path, 'w', encoding='utf-8') as f:
    for line in manifest:
        f.write(line + '\n')

# --- Summary ---
print(f'\nFiles processed: {len(all_files) - len(unsupported)}')
print(f'Outputs created: {total_outputs}')
print(f'Output directory: {outdir}')
print(f'Manifest: {manifest_path}')
if unsupported:
    print(f'\nUNSUPPORTED ({len(unsupported)}):')
    for f in unsupported:
        print(f'  {f}')
if errors:
    print(f'\nERRORS ({len(errors)}):')
    for err in errors:
        print(f'  {err}')
